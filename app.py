import io
import os
import zipfile
import inspect
import tempfile
import shutil
import pandas as pd
import openpyxl
import xml.etree.ElementTree as ET
import streamlit as st

# Monkeypatch openpyxl ColumnDimension & RowDimension to ignore unexpected attributes
# (like 'phonetic') that can be output by third-party tools like Jedox.
import openpyxl.worksheet.dimensions as dim

original_init = dim.ColumnDimension.__init__
def patched_init(self, *args, **kwargs):
    sig = inspect.signature(original_init)
    filtered_kwargs = {k: v for k, v in kwargs.items() if k in sig.parameters.keys()}
    original_init(self, *args, **filtered_kwargs)
dim.ColumnDimension.__init__ = patched_init

original_row_init = dim.RowDimension.__init__
def patched_row_init(self, *args, **kwargs):
    sig = inspect.signature(original_row_init)
    filtered_kwargs = {k: v for k, v in kwargs.items() if k in sig.parameters.keys()}
    original_row_init(self, *args, **filtered_kwargs)
dim.RowDimension.__init__ = patched_row_init

# Set Streamlit Page Configuration
st.set_page_config(
    page_title="Excel Sheet Consolidator",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Inject Premium Glassmorphism CSS Styles & Keyframe Animations
st.markdown("""
<style>
    /* Import Google Font */
    @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800&display=swap');
    
    /* Apply globally */
    html, body, [data-testid="stAppViewContainer"], .stApp {
        font-family: 'Outfit', sans-serif !important;
        background: linear-gradient(135deg, #09070f 0%, #110c22 45%, #050409 100%) !important;
        color: #e2e8f0 !important;
    }
    
    /* Headers styling */
    h1, h2, h3, h4, h5, h6 {
        font-family: 'Outfit', sans-serif !important;
        font-weight: 700 !important;
        color: #ffffff !important;
    }
    
    /* Glassmorphic Cards with Fade-in slide animation */
    .glass-card {
        background: rgba(255, 255, 255, 0.03);
        backdrop-filter: blur(12px);
        -webkit-backdrop-filter: blur(12px);
        border: 1px solid rgba(255, 255, 255, 0.08);
        border-radius: 16px;
        padding: 24px;
        box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.4);
        margin-bottom: 20px;
        animation: slideUp 0.6s cubic-bezier(0.16, 1, 0.3, 1) forwards;
    }
    
    /* Animation Keyframes */
    @keyframes slideUp {
        0% {
            opacity: 0;
            transform: translateY(24px);
        }
        100% {
            opacity: 1;
            transform: translateY(0);
        }
    }
    
    @keyframes fadeIn {
        from { opacity: 0; }
        to { opacity: 1; }
    }
    
    @keyframes float {
        0% { transform: translateY(0px) rotate(0deg); }
        50% { transform: translateY(-8px) rotate(2deg); }
        100% { transform: translateY(0px) rotate(0deg); }
    }
    
    /* Dynamic File Cards */
    .file-card {
        background: rgba(139, 92, 246, 0.06);
        border-left: 4px solid #8b5cf6;
        border-radius: 8px;
        padding: 12px 16px;
        margin: 8px 0;
        display: flex;
        align-items: center;
        justify-content: space-between;
        transition: all 0.3s ease;
        animation: fadeIn 0.4s ease-out forwards;
    }
    .file-card:hover {
        background: rgba(139, 92, 246, 0.12);
        transform: translateX(6px);
        box-shadow: 0 4px 12px rgba(139, 92, 246, 0.1);
    }
    
    /* Title Gradient styling */
    .title-gradient {
        background: linear-gradient(90deg, #a78bfa 0%, #f472b6 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        font-size: 2.0rem !important;
        font-weight: 800 !important;
        margin-bottom: 0.1rem;
        animation: slideUp 0.8s cubic-bezier(0.16, 1, 0.3, 1) forwards;
    }
    
    .subtitle {
        color: #94a3b8;
        font-size: 0.95rem;
        margin-bottom: 1.0rem;
        animation: slideUp 1s cubic-bezier(0.16, 1, 0.3, 1) forwards;
    }
    
    /* Customize Streamlit uploader */
    div[data-testid="stFileUploader"] {
        background-color: rgba(255, 255, 255, 0.02) !important;
        border: 2px dashed rgba(139, 92, 246, 0.3) !important;
        border-radius: 16px !important;
        padding: 20px !important;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
    }
    div[data-testid="stFileUploader"]:hover {
        border-color: #a78bfa !important;
        background-color: rgba(139, 92, 246, 0.04) !important;
        box-shadow: 0 0 15px rgba(139, 92, 246, 0.15);
    }
    
    /* Customize Streamlit Dataframe wrapper */
    div[data-testid="stDataFrame"] {
        border-radius: 12px !important;
        overflow: hidden;
        border: 1px solid rgba(255, 255, 255, 0.08) !important;
        background: rgba(15, 12, 30, 0.6);
    }
    
    /* Custom buttons */
    div.stButton > button, div.stDownloadButton > button {
        background: linear-gradient(90deg, #7c3aed 0%, #db2777 100%) !important;
        color: white !important;
        border: none !important;
        border-radius: 10px !important;
        padding: 10px 24px !important;
        font-weight: 600 !important;
        font-size: 1rem !important;
        box-shadow: 0 4px 14px rgba(124, 58, 237, 0.4) !important;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
    }
    
    div.stButton > button:hover, div.stDownloadButton > button:hover {
        transform: translateY(-2px) scale(1.02) !important;
        box-shadow: 0 6px 20px rgba(219, 39, 119, 0.6) !important;
        background: linear-gradient(90deg, #6d28d9 0%, #be185d 100%) !important;
    }

    /* Floating illustration/icon */
    .floating-logo {
        font-size: 2.2rem;
        text-align: center;
        animation: float 4s ease-in-out infinite;
        margin-bottom: 0.3rem;
    }
    
    /* Badge styling */
    .badge {
        background: rgba(139, 92, 246, 0.15);
        color: #d8b4fe;
        padding: 4px 10px;
        border-radius: 9999px;
        font-size: 0.8rem;
        font-weight: 600;
        border: 1px solid rgba(139, 92, 246, 0.25);
    }
</style>
""", unsafe_allow_html=True)

# App Title Section
st.markdown("""
<div style="text-align: center; margin-top: 0.2rem;">
    <div class="floating-logo">📊</div>
    <div class="title-gradient">Excel Sheet Consolidator</div>
    <div class="subtitle">Extract and combine multiple Excel sheets from a ZIP archive into a single workbook</div>
</div>
""", unsafe_allow_html=True)

# Consolidation Settings (Defaulting directly to Separate Tabs preserving all formatting)
merge_mode = "📂 Separate Tabs (Preserve Format)"
header_row = None
clean_empty = False
add_file_col = False
add_sheet_col = False

# Pure Python XML/ZIP Container-level Sheet Consolidation Engine
def merge_excel_sheets_zip_level(extracted_paths, clean_names):
    xl_zips = []
    try:
        for f in extracted_paths:
            xl_zips.append(zipfile.ZipFile(f))
        return _merge_excel_sheets_zip_level_impl(xl_zips, clean_names)
    finally:
        for xl_zip in xl_zips:
            try:
                xl_zip.close()
            except Exception:
                pass

def _merge_excel_sheets_zip_level_impl(xl_zips, clean_names):
    # Setup namespaces for Excel XML
    ns = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    ET.register_namespace("", "http://schemas.openxmlformats.org/spreadsheetml/2006/main")
    ET.register_namespace("r", "http://schemas.openxmlformats.org/officeDocument/2006/relationships")

    
    # 1. Merge Shared String Tables to prevent index collisions
    unique_strings = []
    string_to_index = {}
    index_maps = []
    
    for xl_zip in xl_zips:
        file_map = {}
        if "xl/sharedStrings.xml" in xl_zip.namelist():
            xml_data = xl_zip.read("xl/sharedStrings.xml")
            root = ET.fromstring(xml_data)
            si_elements = root.findall(".//ns:si", ns)
            for old_idx, si in enumerate(si_elements):
                si_xml = ET.tostring(si)
                if si_xml not in string_to_index:
                    string_to_index[si_xml] = len(unique_strings)
                    unique_strings.append(si_xml)
                file_map[old_idx] = string_to_index[si_xml]
        index_maps.append(file_map)
        
    # Generate unified sharedStrings.xml content
    shared_strings_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="{len(unique_strings)}" uniqueCount="{len(unique_strings)}">'
    )
    for si_xml in unique_strings:
        shared_strings_xml += si_xml.decode('utf-8')
    shared_strings_xml += '</sst>'
    shared_strings_bytes = shared_strings_xml.encode('utf-8')
    
    # Create output workbook bytes
    out_buf = io.BytesIO()
    dest_zip = zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED)
    
    # Copy global workbook metadata from base template (first file)
    base_zip = xl_zips[0]
    exclude_files = {
        "[Content_Types].xml",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
        "xl/sharedStrings.xml"
    }
    
    for name in base_zip.namelist():
        if name in exclude_files:
            continue
        # Exclude worksheet-specific, drawing-specific, media, embedding, and customProperty files
        if (name.startswith("xl/worksheets/") or name.startswith("xl/drawings/") or 
            name.startswith("xl/media/") or name.startswith("xl/embeddings/") or 
            name.startswith("xl/customProperty")):
            continue
        dest_zip.writestr(name, base_zip.read(name))
        
    # Dynamically copy worksheets, media, and rels
    content_types_override = []
    workbook_sheets = []
    workbook_rels = []
    defined_names_xml_parts = []
    
    for idx, (xl_zip, clean_name) in enumerate(zip(xl_zips, clean_names)):
        sheet_num = idx + 1
        sheet_rel_id = f"rId{sheet_num}"
        
        # 2. Write worksheet file
        ws_xml_data = xl_zip.read("xl/worksheets/sheet1.xml")
        
        # Remap shared string indices
        file_map = index_maps[idx]
        if file_map:
            root = ET.fromstring(ws_xml_data)
            cells = root.findall(".//ns:c[@t='s']", ns)
            for cell in cells:
                v_elem = cell.find("ns:v", ns)
                if v_elem is not None:
                    old_idx = int(v_elem.text)
                    v_elem.text = str(file_map.get(old_idx, old_idx))
            ws_xml_data = ET.tostring(root, encoding="utf-8")
            
        dest_sheet_path = f"xl/worksheets/sheet{sheet_num}.xml"
        dest_zip.writestr(dest_sheet_path, ws_xml_data)
        
        content_types_override.append(
            f'<Override PartName="/{dest_sheet_path}" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        )
        workbook_sheets.append(
            f'<sheet name="{clean_name}" sheetId="{sheet_num}" r:id="{sheet_rel_id}"/>'
        )
        workbook_rels.append(
            f'<Relationship Id="{sheet_rel_id}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet{sheet_num}.xml"/>'
        )
        
        # 3. Copy worksheet relationships
        if "xl/worksheets/_rels/sheet1.xml.rels" in xl_zip.namelist():
            rels_data = xl_zip.read("xl/worksheets/_rels/sheet1.xml.rels").decode('utf-8')
            rels_data = rels_data.replace("drawings/drawing1.xml", f"drawings/drawing{sheet_num}.xml")
            rels_data = rels_data.replace("drawings/vmlDrawing1.vml", f"drawings/vmlDrawing{sheet_num}.vml")
            rels_data = rels_data.replace("embeddings/embeddedObjectsheet1.xml", f"embeddings/embeddedObjectsheet{sheet_num}.xml")
            rels_data = rels_data.replace("customProperty1.bin", f"customProperty{sheet_num}.bin")
            dest_zip.writestr(f"xl/worksheets/_rels/sheet{sheet_num}.xml.rels", rels_data.encode('utf-8'))
            
        # 4. Copy drawing file
        if "xl/drawings/drawing1.xml" in xl_zip.namelist():
            drawing_data = xl_zip.read("xl/drawings/drawing1.xml")
            dest_zip.writestr(f"xl/drawings/drawing{sheet_num}.xml", drawing_data)
            content_types_override.append(
                f'<Override PartName="/xl/drawings/drawing{sheet_num}.xml" ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
            )
            
        # 5. Copy drawing relationships & media
        if "xl/drawings/_rels/drawing1.xml.rels" in xl_zip.namelist():
            d_rels_data = xl_zip.read("xl/drawings/_rels/drawing1.xml.rels").decode('utf-8')
            img_name = "image1.png"
            img_ext = ".png"
            for name in xl_zip.namelist():
                if name.startswith("xl/media/image"):
                    img_name = name.split("/")[-1]
                    img_ext = os.path.splitext(img_name)[1]
                    dest_zip.writestr(f"xl/media/image{sheet_num}{img_ext}", xl_zip.read(name))
                    content_types_override.append(
                        f'<Default Extension="{img_ext[1:]}" ContentType="image/{img_ext[1:]}"/>'
                    )
            d_rels_data = d_rels_data.replace(img_name, f"image{sheet_num}{img_ext}")
            dest_zip.writestr(f"xl/drawings/_rels/drawing{sheet_num}.xml.rels", d_rels_data.encode('utf-8'))
            
        # 6. Copy VML drawing (contains Buttons)
        if "xl/drawings/vmlDrawing1.vml" in xl_zip.namelist():
            vml_data = xl_zip.read("xl/drawings/vmlDrawing1.vml")
            dest_zip.writestr(f"xl/drawings/vmlDrawing{sheet_num}.vml", vml_data)
            content_types_override.append(
                f'<Override PartName="/xl/drawings/vmlDrawing{sheet_num}.vml" ContentType="application/vnd.openxmlformats-officedocument.vmlDrawing"/>'
            )
            
        # 7. Copy OLE embeddings
        if "xl/embeddings/embeddedObjectsheet1.xml" in xl_zip.namelist():
            embed_data = xl_zip.read("xl/embeddings/embeddedObjectsheet1.xml")
            dest_zip.writestr(f"xl/embeddings/embeddedObjectsheet{sheet_num}.xml", embed_data)
            content_types_override.append(
                f'<Override PartName="/xl/embeddings/embeddedObjectsheet{sheet_num}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            )
            
        # 8. Extract defined names
        if "xl/workbook.xml" in xl_zip.namelist():
            wb_data = xl_zip.read("xl/workbook.xml")
            wb_root = ET.fromstring(wb_data)
            dn_elements = wb_root.findall(".//ns:definedName", ns)
            for dn in dn_elements:
                val = dn.text or ""
                new_val = val.replace("'Balance Sheet Comparison'", f"'{clean_name}'").replace("Balance Sheet Comparison", f"'{clean_name}'")
                
                dn_attrs = []
                for k, v in dn.attrib.items():
                    if k == "localSheetId":
                        dn_attrs.append(f'localSheetId="{idx}"')
                    else:
                        dn_attrs.append(f'{k}="{v}"')
                if "localSheetId" not in dn.attrib:
                    dn_attrs.append(f'localSheetId="{idx}"')
                    
                attrs_str = " ".join(dn_attrs)
                defined_names_xml_parts.append(
                    f'<definedName {attrs_str}>{new_val}</definedName>'
                )
                
        # 9. Copy customProperty file
        if "xl/customProperty1.bin" in xl_zip.namelist():
            cp_data = xl_zip.read("xl/customProperty1.bin")
            dest_zip.writestr(f"xl/customProperty{sheet_num}.bin", cp_data)
            content_types_override.append(
                f'<Override PartName="/xl/customProperty{sheet_num}.bin" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.customProperty"/>'
            )
                
    # Write merged sharedStrings.xml
    dest_zip.writestr("xl/sharedStrings.xml", shared_strings_bytes)
    content_types_override.append(
        '<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
    )
    
    # 9. Build and write workbook.xml
    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<workbookPr/>'
        '<bookViews>'
        '<workbookView activeTab="0"/>'
        '</bookViews>'
        '<sheets>'
    ) + "\n".join(workbook_sheets) + '</sheets>'
    
    if defined_names_xml_parts:
        workbook_xml += '<definedNames>' + "\n".join(defined_names_xml_parts) + '</definedNames>'
        
    workbook_xml += '<calcPr calcId="124519" fullCalcOnLoad="1"/></workbook>'
    dest_zip.writestr("xl/workbook.xml", workbook_xml.encode('utf-8'))
    
    # 10. Build and write xl/_rels/workbook.xml.rels
    workbook_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<Relationships xmlns="http://schemas.openxmlformats.org/relationships">'
    ) + "\n".join(workbook_rels) + (
        '<Relationship Id="rId_styles" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
        '<Relationship Id="rId_theme" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/theme" Target="theme/theme1.xml"/>'
        '<Relationship Id="rId_sharedStrings" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/>'
        '</Relationships>'
    )
    dest_zip.writestr("xl/_rels/workbook.xml.rels", workbook_rels_xml.encode('utf-8'))
    
    # 11. Build and write [Content_Types].xml
    content_types_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
        '<Override PartName="/xl/theme/theme1.xml" ContentType="application/vnd.openxmlformats-officedocument.theme+xml"/>'
    ) + "\n".join(set(content_types_override)) + '</Types>'
    dest_zip.writestr("[Content_Types].xml", content_types_xml.encode('utf-8'))
    
    # Close dest_zip as it's finished writing
    dest_zip.close()
    
    return out_buf.getvalue()

# Helper function to dynamically generate unique and clean sheet names based on file name pattern
def generate_sheet_names(excel_files):
    bases = []
    for f in excel_files:
        fname = f.split('/')[-1]
        base = os.path.splitext(fname)[0]
        bases.append(base)
        
    split_names = [base.split('_') for base in bases]
    max_parts = max(len(s) for s in split_names) if split_names else 0
    
    padded_splits = []
    for s in split_names:
        padded_splits.append(s + [""] * (max_parts - len(s)))
        
    differing_indices = []
    for i in range(max_parts):
        words_at_i = [s[i] for s in padded_splits]
        if len(set(words_at_i)) > 1:
            differing_indices.append(i)
            
    if not differing_indices:
        chosen_index = max_parts - 1 if max_parts > 0 else 0
    else:
        chosen_index = differing_indices[-1]
        
    sheet_names = []
    for idx, s in enumerate(padded_splits):
        word = s[chosen_index] if chosen_index < len(s) else ""
        if not word:
            word = f"Sheet{idx+1}"
            
        for char in r"\/?:*[]":
            word = word.replace(char, "")
            
        clean_name = word[:31]
        sheet_names.append(clean_name)
        
    seen = {}
    final_names = []
    for name in sheet_names:
        if not name:
            name = "Sheet"
        orig_name = name
        counter = 1
        while name in seen or len(name) == 0:
            suffix = f"_{counter}"
            name = orig_name[:31 - len(suffix)] + suffix
            counter += 1
        seen[name] = True
        final_names.append(name)
        
    return final_names

# Helper function to dynamically generate output Excel file name using all unchanged index words
def generate_output_filename(excel_files):
    if not excel_files:
        return "consolidated_sheets_formatted.xlsx"
        
    bases = []
    for f in excel_files:
        fname = f.split('/')[-1]
        base = os.path.splitext(fname)[0]
        bases.append(base)
        
    split_names = [base.split('_') for base in bases]
    max_parts = max(len(s) for s in split_names) if split_names else 0
    
    padded_splits = []
    for s in split_names:
        padded_splits.append(s + [""] * (max_parts - len(s)))
        
    unchanged_words = []
    for i in range(max_parts):
        words_at_i = [s[i] for s in padded_splits]
        # Check if all elements are the same AND not empty (empty padding doesn't count as word)
        if len(set(words_at_i)) == 1 and words_at_i[0] != "":
            unchanged_words.append(words_at_i[0])
            
    # Join with "_"
    if unchanged_words:
        # Strip forbidden characters just in case
        filename = "_".join(unchanged_words)
        for char in r'\/?:*[]"<>|':
            filename = filename.replace(char, "")
        return f"{filename}.xlsx"
    else:
        return "consolidated_sheets_formatted.xlsx"

# Logic to parse the ZIP file and merge Excel sheets
def merge_excel_sheets(zip_data_or_path, merge_mode="📂 Separate Tabs (Preserve Format)", header_row=None, clean_empty=True, add_file_col=False, add_sheet_col=False):
    processed_details = []
    
    # Setup temp directory to extract files
    temp_dir = tempfile.mkdtemp()
    
    try:
        # Determine if it's a bytes object or file path and extract to temp_dir
        if isinstance(zip_data_or_path, (bytes, io.BytesIO)):
            zip_bytes = zip_data_or_path.getvalue() if hasattr(zip_data_or_path, 'getvalue') else zip_data_or_path
            zip_temp_path = os.path.join(temp_dir, "input.zip")
            with open(zip_temp_path, "wb") as f:
                f.write(zip_bytes)
            zip_file_obj = zipfile.ZipFile(zip_temp_path)
        else:
            zip_file_obj = zipfile.ZipFile(zip_data_or_path)
            
        with zip_file_obj as z:
            # Get list of Excel files in the ZIP (excluding Mac metadata or Excel temp files)
            excel_files = [
                f for f in z.namelist() 
                if (f.endswith('.xlsx') or f.endswith('.xls')) 
                and not f.startswith('__MACOSX') 
                and not f.split('/')[-1].startswith('~$')
            ]
            
            if not excel_files:
                return None, [], None, "No valid Excel files (.xlsx or .xls) found in the ZIP archive."
                
            # Sort files alphabetically to ensure sheet 100 comes before 200, 300, and US
            excel_files.sort()
            
            # Generate clean sheet names using the new index-based pattern matching
            clean_names = generate_sheet_names(excel_files)
            
            # Generate output filename from all unchanged index words
            out_filename = generate_output_filename(excel_files)
                
            if merge_mode == "📂 Separate Tabs (Preserve Format)":
                extracted_paths = []
                
                for file_path in excel_files:
                    out_path = z.extract(file_path, temp_dir)
                    extracted_paths.append(out_path)
                    
                # Perform pure Python XML/ZIP Container-level Sheet Consolidation
                consolidated_bytes = merge_excel_sheets_zip_level(extracted_paths, clean_names)
                
                # Load with openpyxl to generate preview DataFrames
                preview_dict = {}
                for fpath, cname in zip(extracted_paths, clean_names):
                    wb_src = None
                    try:
                        wb_src = openpyxl.load_workbook(fpath, read_only=True)
                        ws_src = wb_src.active
                        preview_data = []
                        for row in ws_src.iter_rows(values_only=True):
                            preview_data.append(list(row))
                        preview_dict[cname] = pd.DataFrame(preview_data)
                        
                        processed_details.append({
                            "file": fpath.split(os.sep)[-1],
                            "sheet": ws_src.title,
                            "tab_name": cname,
                            "rows": ws_src.max_row,
                            "cols": ws_src.max_column
                        })
                    except Exception:
                        pass
                    finally:
                        if wb_src is not None:
                            try:
                                wb_src.close()
                            except Exception:
                                pass
                
                return (consolidated_bytes, preview_dict), processed_details, out_filename, None
                
            else:
                # Combined Tab Mode (Values only, using Pandas)
                sheets_dict = {}
                for file_idx, file_path in enumerate(excel_files):
                    file_name = file_path.split('/')[-1]
                    try:
                        with z.open(file_path) as f:
                            with pd.ExcelFile(f, engine='openpyxl') as excel_reader:
                                for sheet in excel_reader.sheet_names:
                                    df = pd.read_excel(excel_reader, sheet_name=sheet, header=header_row)
                                
                                if clean_empty:
                                    df = df.dropna(how='all')
                                    df = df.dropna(how='all', axis=1)
                                    
                                df = df.reset_index(drop=True)
                                
                                if not df.empty:
                                    name = clean_names[file_idx]
                                    
                                    if add_sheet_col:
                                        df.insert(0, 'Source Sheet', sheet)
                                    if add_file_col:
                                        df.insert(0, 'Source File', file_name)
                                        
                                    sheets_dict[name] = df
                                    processed_details.append({
                                        "file": file_name,
                                        "sheet": sheet,
                                        "tab_name": name,
                                        "rows": len(df),
                                        "cols": len(df.columns)
                                    })
                    except Exception as e:
                        st.warning(f"Error reading file '{file_name}': {str(e)}")
                        
                if not sheets_dict:
                    return None, [], None, "No data rows could be extracted."
                    
                try:
                    merged_df = pd.concat(list(sheets_dict.values()), ignore_index=True)
                    return merged_df, processed_details, out_filename, None
                except Exception as e:
                    return None, [], None, f"Error combining datasets: {str(e)}"
    finally:
        # Guarantee that zip_file_obj is closed first
        if 'zip_file_obj' in locals() and zip_file_obj is not None:
            try:
                zip_file_obj.close()
            except Exception:
                pass
        # Guarantee that the temp folder is fully deleted after handling files
        try:
            shutil.rmtree(temp_dir)
        except Exception:
            pass

# Logic to generate Excel download file in bytes
def generate_excel_bytes(data_payload, merge_mode="📂 Separate Tabs (Preserve Format)"):
    if merge_mode == "📂 Separate Tabs (Preserve Format)":
        # data_payload is a tuple of (bytes, preview_dict)
        file_bytes, _ = data_payload
        return file_bytes
    else:
        # Combined Tab mode (data_payload is a single DataFrame)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            data_payload.to_excel(writer, index=False, sheet_name="Merged Data")
            worksheet = writer.sheets['Merged Data']
            worksheet.hide_gridlines(0)
            
            for i, col in enumerate(data_payload.columns):
                column_len = data_payload[col].astype(str).str.len().max()
                column_len = max(column_len, len(str(col))) + 3
                worksheet.set_column(i, i, min(column_len, 50))
                
        return output.getvalue()

# Main Application Workspace
col1, col2 = st.columns([1, 1.2])

uploaded_zip = None

with col1:
    # st.markdown('<div class="glass-card">', unsafe_allow_html=True)
    st.markdown("### 📂 Upload ZIP File")
    st.markdown("Upload your ZIP archive containing the Excel reports below:")
    
    uploaded_zip = st.file_uploader(
        "Choose a ZIP file", 
        type="zip", 
        label_visibility="collapsed"
    )
    
    process_local = False
    
    st.markdown('</div>', unsafe_allow_html=True)

with col2:
    st.markdown('<div class="glass-card">', unsafe_allow_html=True)
    st.markdown("### 📊 Consolidation Summary")
    
    # Track state of which ZIP is selected
    active_zip = None
    source_name = ""
    
    if uploaded_zip is not None:
        active_zip = uploaded_zip
        source_name = uploaded_zip.name
    elif process_local and local_zip_exists:
        active_zip = local_zip_path
        source_name = local_zip_path
        
    if active_zip is not None:
        # Show animated spinner/processing log
        with st.spinner("Analyzing ZIP archive and extracting sheets with formats..."):
            merged_data, file_details, out_filename, err = merge_excel_sheets(
                active_zip,
                merge_mode=merge_mode,
                header_row=header_row,
                clean_empty=clean_empty,
                add_file_col=add_file_col,
                add_sheet_col=add_sheet_col
            )
            
        if err:
            st.error(f"❌ {err}")
        else:
            st.success("🎉 Excel sheets consolidated successfully!")
            
            # Show Metrics
            total_rows = 0
            if merge_mode == "📊 Combined Tab (Values Only)":
                total_rows = len(merged_data)
            else:
                # merged_data is a tuple of (bytes, preview_dict)
                _, preview_dict = merged_data
                total_rows = sum(len(df) for df in preview_dict.values())
                
            st.markdown(f"""
            <div style="display: flex; gap: 15px; margin: 15px 0;">
                <div style="flex:1; background: rgba(255,255,255,0.02); border: 1px solid rgba(255,255,255,0.06); padding: 12px; border-radius: 10px; text-align:center;">
                    <span style="font-size:0.85rem; color:#94a3b8; display:block;">FILES PROCESSED</span>
                    <strong style="font-size:1.6rem; color:#ffffff;">{len(set(d['file'] for d in file_details))}</strong>
                </div>
                <div style="flex:1; background: rgba(255,255,255,0.02); border: 1px solid rgba(255,255,255,0.06); padding: 12px; border-radius: 10px; text-align:center;">
                    <span style="font-size:0.85rem; color:#94a3b8; display:block;">OUTPUT SHEET TABS</span>
                    <strong style="font-size:1.6rem; color:#ffffff;">{len(file_details)}</strong>
                </div>
                <div style="flex:1; background: rgba(255,255,255,0.02); border: 1px solid rgba(255,255,255,0.06); padding: 12px; border-radius: 10px; text-align:center;">
                    <span style="font-size:0.85rem; color:#94a3b8; display:block;">TOTAL DATA ROWS</span>
                    <strong style="font-size:1.6rem; color:#f472b6;">{total_rows}</strong>
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            # Show which engine was used (for user reassurance)
            engine_badge = "<span class='badge' style='background: rgba(16, 185, 129, 0.15); color: #34d399; border: 1px solid rgba(16, 185, 129, 0.25);'>⚡ Zero-Loss ZIP Engine</span>"
            st.markdown(f"<div style='margin-bottom: 12px;'><strong>Processing Engine:</strong> {engine_badge}</div>", unsafe_allow_html=True)
            
            # Create download package
            with st.spinner("Generating beautiful Excel workbook..."):
                excel_bytes = generate_excel_bytes(merged_data, merge_mode=merge_mode)
                
            # Pulsing Download Button (Moved up!)
            st.markdown("<div style='margin-top: 10px; margin-bottom: 20px; text-align: center;'>", unsafe_allow_html=True)
            st.download_button(
                label="📥 Download Consolidated Excel File",
                data=excel_bytes,
                file_name=st.session_state.get('out_filename', out_filename),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            st.markdown("</div>", unsafe_allow_html=True)

            # File breakdown list with animated CSS hover cards
            with st.expander("📄 View Created Tabs Details", expanded=False):
                for idx, detail in enumerate(file_details):
                    st.markdown(f"""
                    <div class="file-card" style="animation-delay: {idx * 0.1}s;">
                        <div>
                            <strong style="color: #f1f5f9;">{detail['file']}</strong>
                            <div style="font-size: 0.85rem; color: #94a3b8;">Output Tab: <code>{detail['tab_name']}</code></div>
                        </div>
                        <span class="badge">{detail['rows']} Rows</span>
                    </div>
                    """, unsafe_allow_html=True)
            
            # Store in session state
            st.session_state['merged_data'] = merged_data
            st.session_state['out_filename'] = out_filename
            
    else:
        st.info("ℹ️ Please upload a ZIP file on the left to begin.")
        
    st.markdown('</div>', unsafe_allow_html=True)

# Preview section removed as requested
